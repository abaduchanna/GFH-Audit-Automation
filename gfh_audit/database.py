"""SQLite/WAL persistence for variances, status rows, employees, districts,
portal reports and the district audit state machine."""
from __future__ import annotations

import shutil
import sqlite3
import threading
from pathlib import Path
from typing import Dict, Iterable, List, Optional

from .models import InventoryStatusRow, VarianceRow
from .textutils import now_text, person_name_key


class VarianceDatabase:
    def __init__(self, db_path: Path):
        self.db_path = Path(db_path)
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self._lock = threading.RLock()
        self._ensure_wal_mode()
        self._backup_on_startup()
        self._init_db()

    # -- infrastructure -----------------------------------------------------
    def _ensure_wal_mode(self) -> None:
        try:
            con = sqlite3.connect(self.db_path, timeout=30)
            try:
                con.execute("PRAGMA journal_mode=WAL;")
                con.commit()
            finally:
                con.close()
        except Exception:
            pass

    def _backup_on_startup(self) -> None:
        try:
            if self.db_path.exists():
                backup = self.db_path.with_suffix(".sqlite3.bak")
                shutil.copy2(self.db_path, backup)
        except Exception:
            pass

    def connect(self) -> sqlite3.Connection:
        con = sqlite3.connect(self.db_path, timeout=30)
        con.row_factory = sqlite3.Row
        con.execute("PRAGMA journal_mode=WAL;")
        con.execute("PRAGMA busy_timeout=5000;")
        return con

    @staticmethod
    def _table_columns(con: sqlite3.Connection, table: str) -> set:
        rows = con.execute(f"PRAGMA table_info({table})").fetchall()
        return {r["name"] for r in rows}

    def _init_db(self) -> None:
        with self._lock, self.connect() as con:
            con.execute(
                """CREATE TABLE IF NOT EXISTS variances (
                        key TEXT PRIMARY KEY,
                        district TEXT, store TEXT, product TEXT, imei TEXT,
                        status TEXT, created_by TEXT, rep_name TEXT,
                        created_date TEXT, document_status TEXT, source_file TEXT,
                        cleared INTEGER DEFAULT 0, sent_count INTEGER DEFAULT 0,
                        last_sent_at TEXT DEFAULT '', cleared_at TEXT DEFAULT '',
                        cleared_via TEXT DEFAULT '', notes TEXT DEFAULT ''
                   )"""
            )
            con.execute(
                """CREATE TABLE IF NOT EXISTS inventory_status (
                        key TEXT PRIMARY KEY, district TEXT, store TEXT,
                        status TEXT, rep_name TEXT, source_file TEXT,
                        sent_count INTEGER DEFAULT 0, last_sent_at TEXT DEFAULT ''
                   )"""
            )
            con.execute(
                """CREATE TABLE IF NOT EXISTS whatsapp_groups (
                        district TEXT PRIMARY KEY, group_name TEXT
                   )"""
            )
            con.execute(
                """CREATE TABLE IF NOT EXISTS employees (
                        name_key TEXT PRIMARY KEY, name TEXT, phone TEXT, district TEXT,
                        updated_at TEXT
                   )"""
            )
            con.execute(
                """CREATE TABLE IF NOT EXISTS district_runs (
                        district TEXT PRIMARY KEY, state TEXT, started_at TEXT,
                        kickoff_sent_at TEXT, variance_posted_at TEXT,
                        reminders_sent INTEGER DEFAULT 0, last_reminder_at TEXT,
                        final_notice_at TEXT, completed_at TEXT,
                        total_variances INTEGER DEFAULT 0, cleared_variances INTEGER DEFAULT 0,
                        last_message TEXT DEFAULT ''
                   )"""
            )
            con.execute(
                """CREATE TABLE IF NOT EXISTS portal_reports (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        district TEXT, report_type TEXT, file_path TEXT,
                        downloaded_at TEXT, source TEXT
                   )"""
            )
            con.execute(
                """CREATE TABLE IF NOT EXISTS audit_events (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        ts TEXT, district TEXT, event TEXT, detail TEXT
                   )"""
            )
            con.execute(
                """CREATE TABLE IF NOT EXISTS processed_images (
                        message_id TEXT PRIMARY KEY,
                        district TEXT, group_name TEXT,
                        imeis TEXT, processed_at TEXT
                   )"""
            )
            for table, columns in (
                ("variances", {"cleared_via"}),
                ("district_runs", {"last_message", "total_variances", "cleared_variances"}),
            ):
                existing = self._table_columns(con, table)
                for column in columns - existing:
                    con.execute(f"ALTER TABLE {table} ADD COLUMN {column} TEXT DEFAULT ''")

    # -- variances ----------------------------------------------------------
    def upsert_rows(self, rows: List[VarianceRow]) -> None:
        with self._lock, self.connect() as con:
            for row in rows:
                existing = con.execute(
                    "SELECT cleared, cleared_at, cleared_via, sent_count, last_sent_at, notes "
                    "FROM variances WHERE key = ?", (row.key,)
                ).fetchone()
                if existing:
                    con.execute(
                        """UPDATE variances SET district=?, store=?, product=?, imei=?,
                            status=?, created_by=?, rep_name=?, created_date=?,
                            document_status=?, source_file=?, notes=? WHERE key=?""",
                        (row.district, row.store, row.product, row.imei, row.status,
                         row.created_by, row.rep_name, row.created_date,
                         row.document_status, row.source_file, row.notes, row.key),
                    )
                else:
                    con.execute(
                        """INSERT INTO variances (key, district, store, product, imei, status,
                            created_by, rep_name, created_date, document_status, source_file,
                            cleared, cleared_via, notes)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (row.key, row.district, row.store, row.product, row.imei, row.status,
                         row.created_by, row.rep_name, row.created_date, row.document_status,
                         row.source_file, int(row.cleared), row.cleared_via, row.notes),
                    )

    def _row_to_variance(self, r: sqlite3.Row) -> VarianceRow:
        return VarianceRow(
            key=r["key"], district=r["district"] or "", store=r["store"] or "",
            product=r["product"] or "", imei=r["imei"] or "", status=r["status"] or "",
            created_by=r["created_by"] or "", rep_name=r["rep_name"] or "",
            created_date=r["created_date"] or "", document_status=r["document_status"] or "",
            source_file=r["source_file"] or "", cleared=bool(r["cleared"]),
            sent_count=int(r["sent_count"] or 0), last_sent_at=r["last_sent_at"] or "",
            cleared_at=r["cleared_at"] or "", cleared_via=r["cleared_via"] or "",
            notes=r["notes"] or "",
        )

    def rows(self, include_cleared: bool = False, district: str = "") -> List[VarianceRow]:
        query = "SELECT * FROM variances"
        clauses, params = [], []
        if not include_cleared:
            clauses.append("cleared = 0")
        if district:
            clauses.append("district = ?")
            params.append(district)
        if clauses:
            query += " WHERE " + " AND ".join(clauses)
        query += " ORDER BY district, store, imei"
        with self._lock, self.connect() as con:
            return [self._row_to_variance(r) for r in con.execute(query, params).fetchall()]

    def get_rows_by_keys(self, keys: Iterable[str]) -> List[VarianceRow]:
        keys = [k for k in keys]
        if not keys:
            return []
        out: List[VarianceRow] = []
        with self._lock, self.connect() as con:
            for i in range(0, len(keys), 400):
                chunk = keys[i:i + 400]
                marks = ",".join("?" for _ in chunk)
                for r in con.execute(
                    f"SELECT * FROM variances WHERE key IN ({marks})", chunk
                ).fetchall():
                    out.append(self._row_to_variance(r))
        return out

    def set_cleared(self, key: str, cleared: bool, via: str = "manual") -> bool:
        with self._lock, self.connect() as con:
            cur = con.execute(
                "UPDATE variances SET cleared=?, cleared_at=?, cleared_via=? WHERE key=?",
                (int(cleared), now_text() if cleared else "", via if cleared else "", key),
            )
            con.commit()
            return cur.rowcount > 0

    def clear_by_imei(self, imei: str, district: str = "", via: str = "ocr") -> List[str]:
        """Mark every pending variance matching this IMEI as cleared.

        Returns the list of variance keys that were cleared. Matching is
        digit-normalised and also tolerates truncated OCR reads by comparing
        the trailing 11+ digits."""
        import re

        wanted = re.sub(r"\D", "", imei or "")
        if len(wanted) < 11:
            return []
        cleared_keys: List[str] = []
        with self._lock, self.connect() as con:
            query = "SELECT key, imei, district FROM variances WHERE cleared = 0"
            params: list = []
            if district:
                query += " AND district = ?"
                params.append(district)
            for r in con.execute(query, params).fetchall():
                candidate = re.sub(r"\D", "", r["imei"] or "")
                if not candidate:
                    continue
                if candidate == wanted or (
                    len(wanted) >= 12 and candidate.endswith(wanted[-12:])
                ) or (len(candidate) >= 12 and wanted.endswith(candidate[-12:])):
                    con.execute(
                        "UPDATE variances SET cleared=1, cleared_at=?, cleared_via=? WHERE key=?",
                        (now_text(), via, r["key"]),
                    )
                    cleared_keys.append(r["key"])
            con.commit()
        return cleared_keys

    def mark_sent(
        self, rows: Iterable[VarianceRow], group_name: str, batch_title: str,
        image_path: str, mode: str, error: str = "",
    ) -> None:
        with self._lock, self.connect() as con:
            for row in rows:
                con.execute(
                    """UPDATE variances SET sent_count = sent_count + 1,
                        last_sent_at = ? WHERE key = ?""",
                    (now_text(), row.key),
                )
            con.commit()

    # -- inventory status ---------------------------------------------------
    def upsert_inventory_status_rows(self, rows: List[InventoryStatusRow]) -> None:
        with self._lock, self.connect() as con:
            for row in rows:
                con.execute(
                    """INSERT INTO inventory_status (key, district, store, status, rep_name,
                            source_file, sent_count, last_sent_at)
                        VALUES (?,?,?,?,?,?,0,'')
                        ON CONFLICT(key) DO UPDATE SET
                            district=excluded.district, store=excluded.store,
                            status=excluded.status, rep_name=excluded.rep_name,
                            source_file=excluded.source_file""",
                    (row.key, row.district, row.store, row.status, row.rep_name, row.source_file),
                )
            con.commit()

    def inventory_status_rows(self, district: str = "") -> List[InventoryStatusRow]:
        query = "SELECT * FROM inventory_status"
        params: list = []
        if district:
            query += " WHERE district = ?"
            params.append(district)
        query += " ORDER BY district, store"
        with self._lock, self.connect() as con:
            return [
                InventoryStatusRow(
                    key=r["key"], district=r["district"] or "", store=r["store"] or "",
                    status=r["status"] or "", rep_name=r["rep_name"] or "",
                    source_file=r["source_file"] or "",
                )
                for r in con.execute(query, params).fetchall()
            ]

    def mark_status_sent(self, rows: Iterable[InventoryStatusRow]) -> None:
        with self._lock, self.connect() as con:
            for row in rows:
                con.execute(
                    """UPDATE inventory_status SET sent_count = sent_count + 1,
                        last_sent_at = ? WHERE key = ?""",
                    (now_text(), row.key),
                )
            con.commit()

    # -- whatsapp groups ----------------------------------------------------
    def save_whatsapp_group(self, district: str, group_name: str) -> None:
        with self._lock, self.connect() as con:
            con.execute(
                "INSERT INTO whatsapp_groups (district, group_name) VALUES (?,?) "
                "ON CONFLICT(district) DO UPDATE SET group_name=excluded.group_name",
                (district, group_name),
            )
            con.commit()

    def delete_whatsapp_group(self, district: str) -> None:
        with self._lock, self.connect() as con:
            con.execute("DELETE FROM whatsapp_groups WHERE district=?", (district,))
            con.commit()

    def whatsapp_groups(self) -> List[Dict[str, str]]:
        with self._lock, self.connect() as con:
            return [
                {"district": r["district"], "group_name": r["group_name"]}
                for r in con.execute(
                    "SELECT district, group_name FROM whatsapp_groups ORDER BY district"
                ).fetchall()
            ]

    def find_whatsapp_group(self, district: str) -> str:
        with self._lock, self.connect() as con:
            r = con.execute(
                "SELECT group_name FROM whatsapp_groups WHERE district=?", (district,)
            ).fetchone()
            return r["group_name"] if r else ""

    # -- employees ----------------------------------------------------------
    def save_employee(self, name: str, phone: str, district: str = "") -> None:
        name_key = person_name_key(name)
        if not name_key:
            return
        with self._lock, self.connect() as con:
            con.execute(
                """INSERT INTO employees (name_key, name, phone, district, updated_at)
                    VALUES (?,?,?,?,?)
                    ON CONFLICT(name_key) DO UPDATE SET
                        name=excluded.name, phone=excluded.phone,
                        district=excluded.district, updated_at=excluded.updated_at""",
                (name_key, name.strip(), phone.strip(), district.strip(), now_text()),
            )
            con.commit()

    def delete_employee(self, name: str) -> None:
        with self._lock, self.connect() as con:
            con.execute("DELETE FROM employees WHERE name_key=?", (person_name_key(name),))
            con.commit()

    def employees(self) -> List[Dict[str, str]]:
        with self._lock, self.connect() as con:
            return [
                {"name": r["name"], "phone": r["phone"] or "", "district": r["district"] or ""}
                for r in con.execute(
                    "SELECT name, phone, district FROM employees ORDER BY name"
                ).fetchall()
            ]

    def employee_phone_map(self) -> Dict[str, str]:
        """person_name_key -> phone."""
        with self._lock, self.connect() as con:
            return {
                person_name_key(r["name"]): (r["phone"] or "")
                for r in con.execute("SELECT name, phone FROM employees").fetchall()
            }

    def find_employee_phone(self, rep_name: str) -> str:
        """Look up an employee phone by (fuzzy) name.

        Strategy: exact name-key match, then token-overlap scoring."""
        target = person_name_key(rep_name)
        if not target:
            return ""
        employees = self.employees()
        for emp in employees:
            if person_name_key(emp["name"]) == target:
                return emp["phone"] or ""
        target_tokens = set(target.split())
        best_name, best_score = "", 0
        for emp in employees:
            key = person_name_key(emp["name"])
            if not key:
                continue
            tokens = set(key.split())
            overlap = len(tokens & target_tokens)
            if overlap > best_score:
                best_name, best_score = emp["phone"] or "", overlap
        return best_name

    # -- district runs / state machine --------------------------------------
    def set_district_state(self, district: str, state: str) -> None:
        with self._lock, self.connect() as con:
            con.execute(
                """INSERT INTO district_runs (district, state, started_at) VALUES (?,?,?)
                    ON CONFLICT(district) DO UPDATE SET state=excluded.state""",
                (district, state, now_text()),
            )
            con.commit()

    def update_district_run(self, district: str, **fields) -> None:
        if not fields:
            return
        allowed = {
            "state", "kickoff_sent_at", "variance_posted_at", "reminders_sent",
            "last_reminder_at", "final_notice_at", "completed_at",
            "total_variances", "cleared_variances", "last_message",
        }
        fields = {k: v for k, v in fields.items() if k in allowed}
        if not fields:
            return
        with self._lock, self.connect() as con:
            con.execute(
                """INSERT INTO district_runs (district, state, started_at) VALUES (?, 'pending', ?)
                    ON CONFLICT(district) DO NOTHING""",
                (district, now_text()),
            )
            sets = ", ".join(f"{k}=?" for k in fields)
            con.execute(
                f"UPDATE district_runs SET {sets} WHERE district=?",
                (*fields.values(), district),
            )
            con.commit()

    def district_run(self, district: str) -> Dict:
        with self._lock, self.connect() as con:
            r = con.execute(
                "SELECT * FROM district_runs WHERE district=?", (district,)
            ).fetchone()
            return dict(r) if r else {}

    def district_runs(self) -> List[Dict]:
        with self._lock, self.connect() as con:
            return [
                dict(r) for r in con.execute("SELECT * FROM district_runs ORDER BY district")
            ]

    def increment_reminders(self, district: str) -> int:
        with self._lock, self.connect() as con:
            con.execute(
                """INSERT INTO district_runs (district, state, started_at, reminders_sent)
                    VALUES (?, 'variance_posted', ?, 1)
                    ON CONFLICT(district) DO UPDATE SET
                        reminders_sent = reminders_sent + 1, last_reminder_at=?""",
                (district, now_text(), now_text()),
            )
            con.commit()
            r = con.execute(
                "SELECT reminders_sent FROM district_runs WHERE district=?", (district,)
            ).fetchone()
            return int(r["reminders_sent"] or 0) if r else 0

    # -- portal reports ------------------------------------------------------
    def save_portal_report(
        self, district: str, report_type: str, file_path: str, source: str
    ) -> None:
        with self._lock, self.connect() as con:
            con.execute(
                """INSERT INTO portal_reports (district, report_type, file_path, downloaded_at, source)
                    VALUES (?,?,?,?,?)""",
                (district, report_type, file_path, now_text(), source),
            )
            con.commit()

    def latest_portal_report(self, district: str, report_type: str) -> str:
        with self._lock, self.connect() as con:
            r = con.execute(
                """SELECT file_path FROM portal_reports
                    WHERE district=? AND report_type=?
                    ORDER BY id DESC LIMIT 1""",
                (district, report_type),
            ).fetchone()
            return r["file_path"] if r else ""

    # -- audit events --------------------------------------------------------
    def log_event(self, district: str, event: str, detail: str = "") -> None:
        with self._lock, self.connect() as con:
            con.execute(
                "INSERT INTO audit_events (ts, district, event, detail) VALUES (?,?,?,?)",
                (now_text(), district, event, detail[:2000]),
            )
            con.commit()

    # -- processed image registry (prevents duplicate IMEI scanning) ----------
    def is_message_processed(self, message_id: str) -> bool:
        if not message_id:
            return False
        with self._lock, self.connect() as con:
            row = con.execute(
                "SELECT 1 FROM processed_images WHERE message_id = ?", (message_id,)
            ).fetchone()
            return bool(row)

    def mark_message_processed(
        self,
        message_id: str,
        district: str = "",
        group_name: str = "",
        imeis: Optional[List[str]] = None,
    ) -> None:
        if not message_id:
            return
        with self._lock, self.connect() as con:
            con.execute(
                """INSERT OR REPLACE INTO processed_images
                   (message_id, district, group_name, imeis, processed_at)
                   VALUES (?,?,?,?,?)""",
                (message_id, district, group_name,
                 ",".join(imeis or []), now_text()),
            )
            con.commit()
        self.prune_processed_images()

    def prune_processed_images(self, keep: int = 5000) -> None:
        """Keep the registry bounded - newest entries (by insertion) win."""
        with self._lock, self.connect() as con:
            con.execute(
                """DELETE FROM processed_images WHERE rowid NOT IN (
                       SELECT rowid FROM processed_images
                       ORDER BY rowid DESC LIMIT ?
                   )""",
                (keep,),
            )
            con.commit()

    def recent_events(self, limit: int = 200) -> List[Dict]:
        with self._lock, self.connect() as con:
            return [
                dict(r)
                for r in con.execute(
                    "SELECT * FROM audit_events ORDER BY id DESC LIMIT ?", (limit,)
                ).fetchall()
            ]

    # -- settings -------------------------------------------------------------
    def get_setting(self, key: str, default: str = "") -> str:
        with self._lock, self.connect() as con:
            con.execute(
                "CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT)"
            )
            r = con.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
            return r["value"] if r and r["value"] is not None else default

    def save_setting(self, key: str, value: str) -> None:
        with self._lock, self.connect() as con:
            con.execute(
                "CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT)"
            )
            con.execute(
                "INSERT INTO settings (key, value) VALUES (?,?) "
                "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                (key, value),
            )
            con.commit()

    # -- export ----------------------------------------------------------------
    def export_xlsx(self, output_path: Path, keys: Optional[Iterable[str]] = None) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        rows = self.rows(include_cleared=True)
        if keys is not None:
            wanted = set(keys)
            rows = [r for r in rows if r.key in wanted]

        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Log"
        headers = [
            "District", "Store", "Product", "IMEI", "Status", "Rep", "Created By",
            "Created Date", "Cleared", "Cleared At", "Cleared Via", "Sent Count",
            "Last Sent", "Source",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
        for r in rows:
            ws.append([
                r.district, r.store, r.product, r.imei, r.status, r.rep_name,
                r.created_by, r.created_date, "Yes" if r.cleared else "No",
                r.cleared_at, r.cleared_via, r.sent_count, r.last_sent_at, r.source_file,
            ])
        for column, width in zip("ABCDEFGHIJKLMNO", (16, 30, 30, 20, 12, 20, 18, 20, 10, 20, 12, 12, 20, 24)):
            ws.column_dimensions[column].width = width
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
